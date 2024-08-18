import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Directory definitions
directory_path = r'C:\Users\SASH723\Documents\לימודים\פרוייקט גמר\תוצאות הניסוי'
graphs_directory = f'C:\\Users\\SASH723\\Documents\\לימודים\\פרוייקט גמר\\ניתוחים נוספים\\Graphs'

# Ensure the graphs directory exists
os.makedirs(graphs_directory, exist_ok=True)

# Normalize sheet names by removing trailing numbers
def normalize_sheet_name(name):
    return re.sub(r'\d+$', '', name)

# Extract participant and angle from the filename
def extract_participant_angle(filename):
    match = re.search(r'Participant (\d+) Angle (\d+)', filename)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None

# Main processing function
def process_and_export_to_excel(directory_path):
    file_paths = [os.path.join(directory_path, f) for f in os.listdir(directory_path) if f.endswith('.xlsx')]
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    # Prepare data structure for angle data
    data_collection = {}

    # Gather data
    for file_path in file_paths:
        participant, chair_angle = extract_participant_angle(os.path.basename(file_path))
        if participant is None or chair_angle is None:
            continue

        excel_data = pd.ExcelFile(file_path)
        for sheet_name in excel_data.sheet_names:
            if 'success' in sheet_name:
                continue
            normalized_name = normalize_sheet_name(sheet_name)
            df = excel_data.parse(sheet_name)
            angle_data = df.iloc[-4:]  # Last 4 rows are the angles

            for index, row in angle_data.iterrows():
                key = (normalized_name, chair_angle, index - len(df) + 5)
                if key not in data_collection:
                    data_collection[key] = []
                data_collection[key].append((participant, row[1:].values))  # Store time series data

    # Plot and save data
    for (exercise, angle, angle_index), series_list in data_collection.items():
        min_length = min(len(series[1]) for series in series_list)  # Find the shortest series
        time_points = np.arange(min_length)  # Adjust time points to match the shortest series

        plt.figure(figsize=(10, 6))
        for participant, series in series_list:
            plt.plot(time_points, series[:min_length], label=f'Participant {participant}')  # Use only the first min_length points

        plt.title(f'{exercise}, Chair Angle {angle}, Angle Index {angle_index}')
        plt.xlabel('Time')
        plt.ylabel('Angle')
        plt.legend()
        plt.grid(True)

        # Save plot as image
        img_filename = f'{exercise}_Angle{angle}_AngleIndex{angle_index}.png'
        img_path = os.path.join(graphs_directory, img_filename)
        plt.savefig(img_path)
        plt.close()

        # Embed in Excel
        sheet_id = f"{exercise} Angle {angle}"
        if sheet_id not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sheet_id)
        else:
            sheet = workbook[sheet_id]
        img = Image(img_path)
        sheet.add_image(img, 'A1')

    # Save the Excel workbook
    workbook.save(os.path.join(directory_path, 'Combined_Angles_Report.xlsx'))

# Run the function
process_and_export_to_excel(directory_path)
