import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Directory settings
directory_path = r'C:\Users\SASH723\Documents\לימודים\פרוייקט גמר\תוצאות הניסוי'
graphs_directory = f'C:\\Users\\SASH723\\Documents\\לימודים\\פרוייקט גמר\\ניתוחים נוספים\\Graphs2'

# Ensure the graphs directory exists
os.makedirs(graphs_directory, exist_ok=True)

# Normalize sheet names by removing trailing numbers
def normalize_sheet_name(name):
    return re.sub(r'\d+$', '', name)

# Extract participant and angle from the filename
def extract_participant_angle(filename):
    match = re.search(r'Participant (\d+) Angle (\d+)', filename)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None

# Main processing function
def process_and_export_to_excel(directory_path):
    file_paths = [os.path.join(directory_path, f) for f in os.listdir(directory_path) if f.endswith('.xlsx')]
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    # Data structure for plotting
    plot_data = {}

    # Collect data
    for file_path in file_paths:
        participant, chair_angle = extract_participant_angle(os.path.basename(file_path))
        if participant is None or chair_angle is None:
            continue

        excel_data = pd.ExcelFile(file_path)
        for sheet_name in excel_data.sheet_names:
            if 'success' in sheet_name:
                continue
            normalized_name = normalize_sheet_name(sheet_name)
            df = excel_data.parse(sheet_name)
            angle_data = df.iloc[-4:]  # Last 4 rows are the angles

            for index, row in angle_data.iterrows():
                key = (participant, normalized_name, index - len(df) + 5)
                if key not in plot_data:
                    plot_data[key] = {}
                plot_data[key][chair_angle] = row[1:].values

    # Plotting data
    for (participant, exercise, angle_index), angles in plot_data.items():
        plt.figure(figsize=(10, 6))
        min_length = min(len(data) for data in angles.values())  # Shortest series
        time_points = np.arange(min_length)  # Common set of time points

        for chair_angle, series in angles.items():
            plt.plot(time_points, series[:min_length], label=f'Chair Angle {chair_angle}')

        plt.title(f'Participant {participant}, Exercise {exercise}, Angle {angle_index}')
        plt.xlabel('Time')
        plt.ylabel('Angle')
        plt.legend()
        plt.grid(True)

        # Save plot as image
        img_filename = f'P{participant}_{exercise}_Angle{angle_index}.png'
        img_path = os.path.join(graphs_directory, img_filename)
        plt.savefig(img_path)
        plt.close()

        # Embed in Excel
        sheet_id = f"{exercise} P{participant} Angle {angle_index}"
        if sheet_id not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sheet_id)
        else:
            sheet = workbook[sheet_id]
        img = Image(img_path)
        sheet.add_image(img, 'A1')

    # Save the workbook
    workbook.save(os.path.join(directory_path, 'Detailed_Angles_Report.xlsx'))

# Call the function to process files and export results
process_and_export_to_excel(directory_path)
